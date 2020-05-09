import os
import sys
import xlrd  # 读取excel
from openpyxl import load_workbook  # 写入excel
from wordcloud import WordCloud, STOPWORDS, ImageColorGenerator
import jieba  # 结巴分词
import matplotlib.pyplot as plt  # 绘图
from collections import defaultdict  # 字典，用于词频统计
from PIL import Image  # 打开图片，用于词云背景层
import numpy as np  # 转换图片，用于词云背景层

# path = '/Users/fengli/Documents/04Personal/04PersonalFile/cipin_python/data/data.xls'  # 要分词的内容路径
savedFrequencyResultExcel = os.path.join(os.path.dirname(__file__), 'frequency_result.xlsx')
savedImg = os.path.join(os.path.dirname(__file__), 'pic', 'wc_pic.png')
planImg = os.path.join(os.path.dirname(__file__), 'pic', 'background.png')
stopWordsFile = os.path.join(os.path.dirname(__file__), 'src', 'stopwords.txt')
font = os.path.join(os.path.dirname(__file__), 'font', 'simfang.ttf')


def write_result_excel(savedExcel, wordfrequency):
    file = load_workbook(savedExcel)  # 打开excel
    nsheet = file.create_sheet('frequency', index=0)  # 新建表
    nsheet.cell(1, 1, 'word')  # 写入表头
    nsheet.cell(1, 2, 'frequency')  # 写入表头

    wordfrequency_order = sorted(wordfrequency.items(), key=lambda x: x[1], reverse=True)  # 把字典按词频降序排列

    for n in range(2, len(wordfrequency_order) + 2):  # 把降序后的词频统计结果写入excel
        nsheet.cell(n, 1, wordfrequency_order[n - 2][0])
        nsheet.cell(n, 2, wordfrequency_order[n - 2][1])

    file.save(savedExcel)


def usage():
    if len(sys.argv) != 4 or sys.argv[1] in {"-h", "--help"}:  # 判断是否有输入参数，如果输入参数为空或者为-h， -help，则输出帮助信息。
        print("USAGE: python excelCount.py excelName sheetName columnNumber")
        sys.exit()


usage()
path = sys.argv[1]
sheetName = sys.argv[2]
columnNumber = int(sys.argv[3])

exfile = xlrd.open_workbook(path)  # 打开excel
# sheet1 = exfile.sheet_by_index(1)
sheet1 = exfile.sheet_by_name(sheetName)  # 读取Sheet1的内容，根据实际情况填写表名

n = sheet1.nrows  # 表的总行数
mytext = ''
for i in range(0, n):
    text = sheet1.row(i)[columnNumber].value  # 从第0行开始计数
    mytext = mytext + " " + str.lower(text)  # 把每一天内容合并到一个str中

# print(mytext)

# jieba.add_word('花加')  # 如果新词多的话，可以用自己的词库进行分词，在搜狗输入法的网站上有很多分类词库
mytext = jieba.lcut(mytext)  # 把分词结果生成为列表
# print(mytext)

sl = []
with open(stopWordsFile,
          'r') as f:  # 打开停用词文件
    s = f.readlines()
    for a in s:
        a = a.replace('\n', '')
        sl.append(a)  # 把停用词存入列表

sl.append('\u200b')
sl.append('\xa0')  # 这2个符号无法通过stopwords去除，只能在这里增加到列表中，不知道有没有更好的办法

wordfrequency = defaultdict(int)
for word in mytext:
    if word not in sl:  # 去停用词
        wordfrequency[word] += 1  # 词频统计

write_result_excel(savedFrequencyResultExcel, wordfrequency)
print("Word frequency result was saved at %s" % savedFrequencyResultExcel)

img = Image.open(planImg)
myimg = np.array(img)  # 转换图片
stopwords = set(STOPWORDS)

wc = WordCloud(background_color="white", max_words=2000, mask=myimg, font_path=font,
               stopwords=stopwords, max_font_size=40, random_state=42)
# generate word cloud
wc.fit_words(wordfrequency)

# # below code could general word cloud with the color same as origianl picture.
# # create coloring from image
# image_colors = ImageColorGenerator(myimg)
# fig, axes = plt.subplots(1, 3)
# axes[0].imshow(wc, interpolation="bilinear")
# # recolor wordcloud and show
# # we could also give color_func=image_colors directly in the constructor
# # axes[1].imshow(wc.recolor(color_func=image_colors), interpolation="bilinear")
# # axes[2].imshow(myimg, cmap=plt.cm.gray, interpolation="bilinear")
# for ax in axes:
#     ax.set_axis_off()
# plt.show()

plt.imshow(wc, interpolation="bilinear")
plt.axis('off')  # 不显示坐标轴
plt.savefig(savedImg, dpi=600)  # 存储图片，dpi就是每英寸里有多少个点，点越多就越清晰。
plt.show()
